# -*- coding:utf-8 -*-
import openpyxl
from openpyxl.styles import Border, Side

class cFileOpt:
	InputCol = ''
	NameCN = ''
	InputFil = 0
	InputSht = 0
	OutputFil = 0
	OutputSht = 0
	bian = Side(style='thin', color='000000')
	border = Border(top=bian, bottom=bian, left=bian, right=bian)
	SrcLst = [0 for col in range(46)]
	DstLst = [0 for col in range(46)]
	def __init__(self):
		self.InputFil = openpyxl.load_workbook('demo.xlsx')
		print('open demo.xlsx success')

	def GetNam(self):
		self.InputCol = input('输入列号，如(G)，代表G列:\n ')
		self.NameCN = input('输入名字，如(狮子王):\n ')
		return self.NameCN

	def CloseFiles(self):
		self.InputFil.close()
		self.OutputFil.close()
		pass

	def Prepare(self):
		self.InputSht = self.InputFil.worksheets[1]
		self.OutputFil = openpyxl.Workbook()
		self.OutputSht = self.OutputFil.active
		self.OutputSht.title = 'Sheet1'
		pass
	
	def SaveOutFiles(self, NameSmp):
		if self.InputCol == 'G':
			Filename = 'Sale_%s.xlsx' % (NameSmp)
		elif self.InputCol == 'H':
			Filename = 'CustServ_%s.xlsx' % (NameSmp)
		else:
			Filename = '%s.xlsx' % (NameSmp)
		self.OutputFil.save(filename = Filename)
		print('%s的数据保存成功！' % NameSmp)
		pass

	def ReadFile(self, wb):
		CntSales = 0
		for i in range(1, self.InputSht.max_row):
			ColCel = ('%s%d' % (self.InputCol, i))
			if self.InputSht[ColCel].value == self.NameCN:
				CntSales += 1
				self.SrcLst[0] = CntSales
				for j in range(1, 46):
					self.SrcLst[j] = self.InputSht.cell(row = i,column = (j + 1)).value
				self.WriteFile(self.SrcLst, CntSales)
		pass

	def WriteFile(self, DstLst, StartLine):
		for i in range(StartLine+1, StartLine+2):
			for j in range(1, 47):
				self.OutputSht.cell(row = i,column = j).value = DstLst[(j - 1)]
				self.OutputSht.cell(row = i,column = j).border = self.border
				self.OutputSht.cell(row = i,column = j).alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
				self.OutputSht.cell(row = i,column = j).font = openpyxl.styles.Font(name='微软雅黑', size=10)

				if j == 40:
					if self.OutputSht.cell(row = i,column = j).value == 0 or self.OutputSht.cell(row = i,column = j).value == '':
						pass
					else:
						self.OutputSht.cell(row = i,column = j).value = ('=IF(D%d="差额发票",ROUND(R%d/1.05*5.77%%,2),IF(D%d="全额发票",ROUND(K%d/1.06*6.77%%,2)))' % (i, i, i, i))
				elif j == 42:
					self.OutputSht.cell(row = i,column = j).value = ('=ROUND(R%d+S%d+T%d+U%d+V%d+W%d+X%d-AE%d-AF%d-AG%d-AH%d-AI%d-AJ%d-AK%d-AM%d-AN%d-AO%d,2)' % (i,i,i,i,i,i,i,i,i,i,i,i,i,i,i,i,i))
				elif j == 43:
					if self.OutputSht.cell(row = i,column = j).value == 0:
						pass
					else:
						self.OutputSht.cell(row = i,column = j).value =('=ROUND(AP%d*20%%,2)' % i)
				elif j == 44:
					self.OutputSht.cell(row = i,column = j).value = ('=AP%d-AQ%d' % (i,i))
				elif j == 46:
					self.OutputSht.cell(row = i,column = j).value = ('=IF((AS%d-I%d)/365<1,"小于1年",IF(AND((AS%d-I%d)/365>=1,(AS%d-I%d)/365<2),"1-2年",IF(AND((AS%d-I%d)/365>=2,(AS%d-I%d)/365<3),"2-3年",IF((AS%d-I%d)/365>=3,"3年以上"," "))))' % (i,i,i,i,i,i,i,i,i,i,i,i))
		pass

	def SetFormat(self):
		for i in range(1, 47):
			self.OutputSht.cell(row = 1,column = i).value = self.InputSht.cell(row = 1,column = i).value
			self.OutputSht.cell(row = 1,column = i).border = self.border
			self.OutputSht.cell(row = 1,column = i).alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
			self.OutputSht.cell(row = 1,column = i).font = openpyxl.styles.Font(name='微软雅黑', size=11, bold=True)
			if i >= 25 and i <= 37:
				self.OutputSht.cell(row = 1,column = i).fill = openpyxl.styles.PatternFill(start_color ='FF0000', end_color = 'FF0000', fill_type = 'solid')
			else:
				self.OutputSht.cell(row = 1,column = i).fill = openpyxl.styles.PatternFill(start_color ='92D050', end_color = '92D050', fill_type = 'solid')
		self.OutputSht.column_dimensions['A'].width = 5.30	# +0.74 with true value(only column width needs)
		self.OutputSht.column_dimensions['B'].width = 49.76
		self.OutputSht.column_dimensions['C'].width = 9.36
		self.OutputSht.column_dimensions['D'].width = 9.58
		self.OutputSht.column_dimensions['E'].width = 30.24
		self.OutputSht.column_dimensions['F'].width = 8.58
		self.OutputSht.column_dimensions['G'].width = 8.58
		self.OutputSht.column_dimensions['I'].width = 20.47
		self.OutputSht.column_dimensions['AS'].width = 20.47
		self.OutputSht.row_dimensions[1].height = 28
		pass

def main():
	FileOpt = cFileOpt()
	Name = FileOpt.GetNam()
	FileOpt.Prepare()
	FileOpt.SetFormat()
	FileOpt.ReadFile(FileOpt.InputFil)
	FileOpt.SaveOutFiles(Name)
	FileOpt.CloseFiles()
	
if __name__ == '__main__':
	main()